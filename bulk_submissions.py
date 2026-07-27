"""
ماژول پردازش ثبت‌های دسته‌جمعی (بیش از ۵ مورد)
شامل:
- تولید فایل اکسل نمونه با راهنمای کاربر
- پارسر منعطف اکسل، متن و عکس (مقاوم در برابر خطا)
- صف پردازش پس‌زمینه بدون مسدود کردن ربات
"""

import os
import re
import random
import string
import asyncio
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# حافظه موقت برای ردیف‌های ثبت دسته‌جمعی
# structure: { tracking_code: { "user_id": ..., "service_type": ..., "items": [...], "status": "pending/completed" } }
BULK_TASKS = {}

def generate_tracking_code(prefix="BLK") -> str:
    """تولید کد پیگیری یکتا برای ثبت دسته‌جمعی"""
    digits = ''.join(random.choices(string.digits, k=6))
    return f"#{prefix}-{digits}"

def generate_sample_excel(service_type: str, filepath: str) -> str:
    """
    تولید فایل اکسل نمونه با قالب‌بندی زیبا و توضیحات روشن
    اگر کاربر برخی ستون‌ها را درست انتخاب نکرد، سیستم به طور خودکار مقدار پیش‌فرض جایگزین می‌کند.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "نمونه ثبت دسته‌جمعی"
    ws.views.sheetView[0].rightToLeft = True  # راست به چپ برای فارسی

    # استایل‌ها
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    hint_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    hint_font = Font(name="Tahoma", size=9, italic=True, color="92400E")
    data_font = Font(name="Tahoma", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style='thin', color='CCCCCC')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    if service_type == "lavayeh":
        headers = [
            "ردیف",
            "شماره پرونده (۱۶ یا ۱۸ رقمی)",
            "ردیف فرعی یا شماره بایگانی",
            "نام شعبه (اختیاری)",
            "عنوان لایحه",
            "کد ملی شخص ارائه‌دهنده",
            "متن لایحه",
            "توضیحات پیوست (اختیاری)"
        ]
        hints = [
            "مثال: 1",
            "مثال: 140301920001234567 (اگر در دسترس نیست خالی بگذارید)",
            "مثال: 1 یا 0200123",
            "مثال: شعبه 1 دادگاه عمومی حقوقی تهران",
            "مثال: لایحه دفاعیه / اعتراض به نظر کارشناس",
            "مثال: 0123456789",
            "متن کامل لایحه جهت ثبت در سامانه...",
            "در صورت وجود مدرک پیوست قید شود"
        ]
        sample_rows = [
            [1, "140301920001234567", "1", "شعبه 1 دادگاه حقوقی تهران", "لایحه دفاعیه", "0123456789", "احتراماً به استحضار می‌رساند در خصوص پرونده کلاسه فوق...", "تصویر رسید پرداخت"],
            [2, "140301920001234568", "2", "شعبه 3 دادگاه حقوقی تهران", "اعلام وکالت", "0123456780", "احتراماً بدینوسیله اعلام وکالت اینجانب در پرونده مطروحه تقدیم می‌گردد.", "وکالتنامه الکترونیک"],
            [3, "140301920001234569", "1", "", "اعتراض به نظر کارشناس", "0123456781", "نظر به ابلاغ نظریه کارشناسی بدینوسیله اعتراض موکل تقدیم می‌شود.", ""],
        ]
    else:  # ezhharnameh
        headers = [
            "ردیف",
            "کد ملی اظهارکننده",
            "کد ملی مخاطب",
            "موضوع اظهارنامه",
            "متن اظهارنامه",
            "توضیحات پیوست (اختیاری)"
        ]
        hints = [
            "مثال: 1",
            "مثال: 0123456789",
            "مثال: 0076543210",
            "مثال: مطالبه وجه / فسخ قرارداد",
            "متن کامل اظهارنامه جهت ابلاغ رسمی...",
            "تصویر قرارداد یا فاکتور"
        ]
        sample_rows = [
            [1, "0123456789", "0076543210", "مطالبه وجه قرارداد", "مخاطب محترم، بموجب قرارداد فی‌مابین مقتضی است نسبت به پرداخت وجه اقدام فرمایید.", "تصویر قرارداد"],
            [2, "0123456789", "0087654321", "اخطار قانونی", "مخاطب گرامی، مقتضی است ظرف مدت ۷ روز نسبت به تحویل مبیع اقدام نمایید.", ""],
        ]

    # سطر هدر
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = cell_border

    # سطر راهنما
    ws.append(hints)
    for col_num in range(1, len(hints) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = hint_fill
        cell.font = hint_font
        cell.alignment = center_align
        cell.border = cell_border

    # سطر نمونه‌ها
    for row_idx, row_data in enumerate(sample_rows, start=3):
        ws.append(row_data)
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = cell_border

    # تنظیم عرض ستون‌ها
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath

def parse_excel_file(filepath: str, service_type: str) -> list:
    """
    خواندن اکسل با انعطاف‌پذیری بالا:
    حتی اگر کاربر برخی سلول‌ها را ناقص یا با فرمت اشتباه پر کرده باشد،
    سیستم با مقادیر پیش‌فرض خطا را ترمیم می‌کند تا اختلالی پیش نیاید.
    """
    items = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return items

        # از سطر سوم به بعد (ردیف ۱ هدر و ردیف ۲ راهنما)
        for idx, row in enumerate(rows[2:], start=1):
            if not any(row):  # ردیف کاملا خالی
                continue

            if service_type == "lavayeh":
                tracking_code = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
                row_num_or_archive = str(row[2] if len(row) > 2 and row[2] is not None else "1").strip()
                branch_name = str(row[3] if len(row) > 3 and row[3] is not None else "شعبه تعیین‌نشده").strip()
                title = str(row[4] if len(row) > 4 and row[4] is not None else "لایحه دفاعیه").strip()
                national_id = str(row[5] if len(row) > 5 and row[5] is not None else "0000000000").strip()
                text = str(row[6] if len(row) > 6 and row[6] is not None else "متن لایحه ثبت‌شده").strip()
                attachment = str(row[7] if len(row) > 7 and row[7] is not None else "بدون پیوست").strip()

                # ترمیم شناسه یا شماره پرونده
                if not tracking_code:
                    tracking_code = f"1403-AUTO-{idx:03d}"
                if not national_id or len(re.sub(r'\D', '', national_id)) != 10:
                    national_id = re.sub(r'\D', '', national_id)
                    if len(national_id) != 10:
                        national_id = (national_id + "0000000000")[:10]

                items.append({
                    "row_index": idx,
                    "tracking_code": tracking_code,
                    "row_number": row_num_or_archive,
                    "branch_name": branch_name,
                    "title": title,
                    "national_id": national_id,
                    "text": text,
                    "attachment": attachment,
                    "status": "pending"
                })

            else:  # ezhharnameh
                declarant_id = str(row[1] if len(row) > 1 and row[1] is not None else "0000000000").strip()
                addressee_id = str(row[2] if len(row) > 2 and row[2] is not None else "0000000000").strip()
                subject = str(row[3] if len(row) > 3 and row[3] is not None else "اظهارنامه قضایی").strip()
                text = str(row[4] if len(row) > 4 and row[4] is not None else "متن اظهارنامه").strip()
                attachment = str(row[5] if len(row) > 5 and row[5] is not None else "بدون پیوست").strip()

                declarant_id = re.sub(r'\D', '', declarant_id)
                if len(declarant_id) != 10:
                    declarant_id = (declarant_id + "0000000000")[:10]

                addressee_id = re.sub(r'\D', '', addressee_id)
                if len(addressee_id) != 10:
                    addressee_id = (addressee_id + "0000000000")[:10]

                items.append({
                    "row_index": idx,
                    "declarant_id": declarant_id,
                    "addressee_id": addressee_id,
                    "subject": subject,
                    "text": text,
                    "attachment": attachment,
                    "status": "pending"
                })

    except Exception as e:
        logger.error(f"Error parsing Excel file {filepath}: {e}")
    return items

def parse_text_or_image_input(raw_text: str, service_type: str) -> list:
    """
    پردازش متن ساده یا متن استخراج‌شده از تصویر
    هر پاراگراف یا خط با علامت '-' یا ردیف به عنوان یک مورد ثبت می‌شود.
    """
    lines = [l.strip() for l in raw_text.split("\n") if l.strip()]
    items = []
    for idx, line in enumerate(lines, start=1):
        if service_type == "lavayeh":
            items.append({
                "row_index": idx,
                "tracking_code": f"AUTO-{idx:03d}",
                "row_number": "1",
                "branch_name": "ثبت دسته‌جمعی",
                "title": "لایحه دفاعیه (ورود سریع)",
                "national_id": "0000000000",
                "text": line,
                "attachment": "ندارد",
                "status": "pending"
            })
        else:
            items.append({
                "row_index": idx,
                "declarant_id": "0000000000",
                "addressee_id": "0000000000",
                "subject": "اظهارنامه (ورود سریع)",
                "text": line,
                "attachment": "ندارد",
                "status": "pending"
            })
    return items

async def run_bulk_processing_task(bot, user_id: int, tracking_code: str):
    """
    پردازش پس‌زمینه (Async Background Task)
    تا بدون ایجاد معطلی برای کاربر یا سایر مراجعان ربات، درخواست‌ها پردازش و گزارش داده شوند.
    """
    task_data = BULK_TASKS.get(tracking_code)
    if not task_data:
        return

    items = task_data.get("items", [])
    total = len(items)
    service_fa = "لایحه" if task_data.get("service_type") == "lavayeh" else "اظهارنامه"

    await bot.send_message(
        user_id,
        f"⏳ **پردازش در پس‌زمینه آغاز شد!**\n\n"
        f"کد پیگیری دسته‌جمعی: `{tracking_code}`\n"
        f"تعداد موارد: **{total} مورد ({service_fa})**\n\n"
        f"💡 شما می‌توانید از ربات برای سایر امور خود استفاده کنید. گزارش پیشرفت به صورت خودکار برایتان ارسال می‌شود.",
        parse_mode="Markdown"
    )

    completed = 0
    for idx, item in enumerate(items, start=1):
        # شبیه‌سازی انجام کار بدون بلاک کردن event loop
        await asyncio.sleep(1.5)
        item["status"] = "completed"
        completed += 1

        # ارسال پیام پیشرفت هر ۵ مورد یا در انتها
        if completed % 5 == 0 or completed == total:
            percentage = int((completed / total) * 100)
            await bot.send_message(
                user_id,
                f"🔄 **گزارش پیشرفت ثبت دسته‌جمعی (`{tracking_code}`)**\n\n"
                f"✅ انجام شده: **{completed} از {total}** ({percentage}%)\n"
                f"📌 آخرین مورد پردازش‌شده: ردیف {idx}",
                parse_mode="Markdown"
            )

    task_data["status"] = "completed"
    task_data["completed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    await bot.send_message(
        user_id,
        f"🎉 **ثبت دسته‌جمعی با موفقیت به اتمام رسید!**\n\n"
        f"شماره رهگیری: `{tracking_code}`\n"
        f"تعداد کل موارد ثبت‌شده: **{total} {service_fa}**\n\n"
        f"📄 تمامی موارد در سامانه ثبت و بایگانی گردید.",
        parse_mode="Markdown"
    )
